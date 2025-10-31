import importlib, pathlib
from openpyxl import load_workbook
from gunsonu_test_adapter import init_db
import db_model

def test_branch_pivot_daily(tmp_path):
    db = tmp_path / "gunsonu.sqlite"
    init_db(db)
    db_model.ensure_schema(str(db))

    # Seed incomes across days
    db_model.insert_income(str(db), date="2025-10-01", amount=100.0, description="Satis", branch="Merkez")
    db_model.insert_income(str(db), date="2025-10-01", amount=50.0,  description="Satis", branch="SubeA")
    db_model.insert_income(str(db), date="2025-10-02", amount=25.5,  description="Satis", branch="Merkez")
    db_model.insert_income(str(db), date="2025-10-15", amount=10.0,  description="Satis", branch="SubeA")

    exporter = importlib.import_module("exporter")
    out = tmp_path / "rapor.xlsx"
    exporter.export_all(str(db), str(out), ym="2025-10")

    wb = load_workbook(out)
    wsp = wb["Sube_Pivot"]
    # header checks
    assert wsp["A1"].value == "Şube"
    assert wsp["C1"].value == "02"  # day padding
    assert wsp.cell(row=1, column=1 + 31 + 1).value == "Toplam"  # last header

    # Find rows by branch
    rows = {r[0]: r for r in wsp.iter_rows(min_row=2, values_only=True)}
    # Merkez: d1=100, d2=25.5, toplam=125.5
    merkez = rows["Merkez"]
    assert abs(merkez[1] - 100.0) < 1e-6
    assert abs(merkez[2] - 25.5) < 1e-6
    assert abs(merkez[-1] - 125.5) < 1e-6

    # SubeA: d1=50, d15=10, toplam=60
    subea = rows["SubeA"]
    assert abs(subea[1] - 50.0) < 1e-6
    # day 15 column index = 1 (branch col) + 15
    assert abs(subea[1 + 15] - 10.0) < 1e-6
    assert abs(subea[-1] - 60.0) < 1e-6