import importlib, pathlib
from openpyxl import load_workbook
from gunsonu_test_adapter import init_db
import db_model

def test_exporter_monthly_summary_and_branch(tmp_path):
    db = tmp_path / "gunsonu.sqlite"
    init_db(db)
    db_model.ensure_schema(str(db))

    # Seed incomes/expenses with branches in 2025-10
    db_model.insert_income(str(db), date="2025-10-02", amount=1000.0, description="Satis", branch="Merkez")
    db_model.insert_income(str(db), date="2025-10-15", amount=500.0, description="Online", branch="Merkez")
    db_model.insert_expense(str(db), date="2025-10-05", amount=300.0, description="Elektrik", branch="Merkez")
    # extra other month (should not count)
    db_model.insert_income(str(db), date="2025-09-30", amount=999.0, description="prev", branch="Merkez")

    # advances for month
    db_model.insert_advance(str(db), person="Ali", date="2025-10-07", amount=200.0, note="")

    exporter = importlib.import_module("exporter")
    out = tmp_path / "rapor.xlsx"
    exporter.export_all(str(db), str(out), ym="2025-10")
    assert out.exists()

    wb = load_workbook(out)

    # Ozet_Aylik checks
    ws = wb["Ozet_Aylik"]
    assert ws["A1"].value == "Ay" and ws["B1"].value == "2025-10"
    labels = [ws[f"A{i}"].value for i in range(2,6)]
    assert labels == ["Toplam Gelir","Toplam Gider","Toplam Avans","Net"]
    vals = [ws[f"B{i}"].value for i in range(2,6)]
    assert vals[0] == 1500.0  # gelir
    assert vals[1] == 300.0   # gider
    assert vals[2] == 200.0   # avans
    assert vals[3] == 1000.0  # net

    # Sube_Kirilim checks
    bk = wb["Sube_Kirilim"]
    assert bk["A1"].value == "Åube"
    assert bk["B1"].value == "Gelir"
    assert bk["C1"].value == "Gider"
    # find Merkez row
    rows = list(bk.iter_rows(min_row=2, values_only=True))
    assert any(r[0] == "Merkez" and r[1] == 1500.0 and r[2] == 300.0 for r in rows)

    # Sheets order (B option): Ozet_Aylik first
    assert wb.sheetnames[:2] == ["Ozet_Aylik","Sube_Kirilim"]
