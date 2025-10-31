import importlib, pathlib
from openpyxl import load_workbook
from gunsonu_test_adapter import init_db
import db_model

def test_z_report_fill(tmp_path):
    db = tmp_path / "gunsonu.sqlite"
    init_db(db)
    db_model.ensure_schema(str(db))

    # two days of Z data for 2025-10
    db_model.upsert_z_report(str(db), date="2025-10-01", cash_total=100.0, pos_total=50.0, qr_included=1)
    db_model.upsert_z_report(str(db), date="2025-10-02", cash_total=200.0, pos_total=25.5, qr_included=0)

    exporter = importlib.import_module("exporter")
    out = tmp_path / "rapor.xlsx"
    exporter.export_all(str(db), str(out), ym="2025-10")
    wb = load_workbook(out)
    ws = wb["Z_Rapor"]

    # header check
    assert ws["A1"].value == "Tarih"
    assert ws["B1"].value == "Kasa Toplam"
    assert ws["C1"].value == "POS Toplam"
    assert ws["D1"].value == "QR_Dahil"

    # rows check
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert ("2025-10-01", 100.0, 50.0, 1) in rows
    assert ("2025-10-02", 200.0, 25.5, 0) in rows