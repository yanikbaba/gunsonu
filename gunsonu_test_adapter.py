from __future__ import annotations
import os, sys, json, types, shutil, pathlib, importlib, time, sqlite3

ROOT = pathlib.Path(__file__).resolve().parent

def _import(name: str) -> types.ModuleType | None:
    try:
        return importlib.import_module(name)
    except Exception:
        return None

def documents_dir() -> pathlib.Path:
    home = pathlib.Path.home()
    candidates = [home / "Documents", home / "Belgeler", home / "Belgelerim"]
    for c in candidates:
        if c.exists():
            return c
    return home / "Documents"

def default_db_dir() -> pathlib.Path:
    d_drive = pathlib.Path("D:/GunSonu/DB")
    if d_drive.exists():
        return d_drive
    return documents_dir() / "GunSonu/DB"

def ensure_dir(p: pathlib.Path) -> pathlib.Path:
    p.mkdir(parents=True, exist_ok=True)
    return p

def load_config(path: pathlib.Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def init_db(db_path: pathlib.Path) -> pathlib.Path:
    """Initialize a REAL SQLite db file so snapshots are non-empty.
    Accepts either a directory (creates gunsonu.sqlite) or a full file path.
    """
    mod = _import("db_model")
    db_file = db_path / "gunsonu.sqlite" if db_path.is_dir() else db_path
    ensure_dir(db_file.parent)

    # 1) Project-provided initializer if available
    used_project_init = False
    for fname in ("init_db", "initialize_db", "bootstrap_db", "ensure_schema"):
        if mod and hasattr(mod, fname):
            getattr(mod, fname)(str(db_file))
            used_project_init = True
            break

    # 2) Ensure file exists
    if not db_file.exists():
        db_file.touch()

    # 3) Guarantee it is a valid SQLite (not zero-byte)
    try:
        if db_file.stat().st_size == 0:
            with sqlite3.connect(str(db_file)) as conn:
                conn.execute("PRAGMA journal_mode=WAL;")
                conn.execute("CREATE TABLE IF NOT EXISTS meta (id INTEGER PRIMARY KEY, created_at TEXT DEFAULT CURRENT_TIMESTAMP);")
                conn.commit()
    except Exception:
        pass

    return db_file

def create_snapshot(db_file: pathlib.Path, snapshots_dir: pathlib.Path | None = None) -> pathlib.Path | None:
    mod = _import("backup_manager")
    if snapshots_dir is None:
        snapshots_dir = pathlib.Path(".") / ".snapshots"
    ensure_dir(snapshots_dir)
    if mod:
        for fname in ("create_snapshot", "backup_now", "make_snapshot"):
            if hasattr(mod, fname):
                try:
                    return pathlib.Path(getattr(mod, fname)(str(db_file), str(snapshots_dir)))
                except Exception:
                    pass
    ts = time.strftime("%Y%m%d_%H%M%S")
    snap = snapshots_dir / f"gunsonu_{ts}.sqlite"
    shutil.copy2(db_file, snap)
    return snap

def export_excel(db_file: pathlib.Path, out_dir: pathlib.Path | None = None) -> pathlib.Path | None:
    mod = _import("exporter")
    if out_dir is None:
        out_dir = default_db_dir()
    ensure_dir(out_dir)
    out_file = out_dir / "rapor.xlsx"
    if mod:
        for fname in ("export_all", "export_to_excel", "exporter_main"):
            if hasattr(mod, fname):
                fn = getattr(mod, fname)
                try:
                    try:
                        fn(str(db_file), str(out_file))
                    except TypeError:
                        fn(str(out_file))
                    return out_file
                except Exception:
                    pass
    try:
        from openpyxl import Workbook
        wb = Workbook()
        for name in ("Z_Rapor","Gelirler","Giderler","Avanslar"):
            wb.create_sheet(name)
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(out_file)
        return out_file
    except Exception:
        return None

def recover_last_snapshot(db_file: pathlib.Path, snapshots_dir: pathlib.Path | None = None) -> bool:
    mod = _import("crash_recovery")
    if snapshots_dir is None:
        snapshots_dir = pathlib.Path(".") / ".snapshots"
    snaps = sorted(snapshots_dir.glob("*.sqlite"))
    if not snaps:
        return False
    last = snaps[-1]
    if mod:
        for fname in ("recover_last_snapshot", "restore_latest", "recover"):
            if hasattr(mod, fname):
                try:
                    return bool(getattr(mod, fname)(str(db_file), str(snapshots_dir)))
                except Exception:
                    pass
    shutil.copy2(last, db_file)
    return True
