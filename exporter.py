# exporter.py â€” Excel export with monthly summary, branch breakdown, Z_Rapor fill
# E3: Sube_Pivot (gÃ¼nlÃ¼k gelir) sayfasÄ± eklendi
from __future__ import annotations
import sqlite3, pathlib, calendar, datetime
from typing import List, Tuple

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

def _connect(db_path: str) -> sqlite3.Connection:
    con = sqlite3.connect(db_path, timeout=15, isolation_level=None)
    con.execute("PRAGMA foreign_keys=ON;")
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=NORMAL;")
    return con

def _monthly_totals(con: sqlite3.Connection, ym: str):
    cur = con.execute("SELECT COALESCE(SUM(amount),0) FROM incomes WHERE substr(date,1,7)=?", (ym,))
    inc = float(cur.fetchone()[0] or 0.0)
    cur = con.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE substr(date,1,7)=?", (ym,))
    exp = float(cur.fetchone()[0] or 0.0)
    cur = con.execute("SELECT COALESCE(SUM(amount),0) FROM advances WHERE substr(date,1,7)=?", (ym,))
    adv = float(cur.fetchone()[0] or 0.0)
    return round(inc,2), round(exp,2), round(adv,2), round(inc-exp-adv,2)

def _branch_breakdown(con: sqlite3.Connection, ym: str) -> List[Tuple[str,float,float]]:
    cur = con.execute(
        """WITH branches AS (
               SELECT DISTINCT branch FROM incomes WHERE substr(date,1,7)=?
               UNION
               SELECT DISTINCT branch FROM expenses WHERE substr(date,1,7)=?
           )
           SELECT b.branch,
                  ROUND(COALESCE((SELECT SUM(amount) FROM incomes  WHERE substr(date,1,7)=? AND branch=b.branch),0),2) AS inc_total,
                  ROUND(COALESCE((SELECT SUM(amount) FROM expenses WHERE substr(date,1,7)=? AND branch=b.branch),0),2) AS exp_total
           FROM branches b
           ORDER BY b.branch COLLATE NOCASE ASC""",
        (ym, ym, ym, ym)
    )
    return [(r[0] or "", float(r[1] or 0.0), float(r[2] or 0.0)) for r in cur.fetchall()]

def _month_len(ym: str) -> int:
    y, m = map(int, ym.split("-"))
    return calendar.monthrange(y, m)[1]

def _branch_pivot_incomes(con: sqlite3.Connection, ym: str):
    # Returns dict: {branch: {day:int -> amount:float}}
    days = _month_len(ym)
    # gather per-branch-per-day sums
    cur = con.execute(
        """SELECT COALESCE(branch,''), CAST(substr(date,9,2) AS INT) AS d, ROUND(SUM(amount),2)
           FROM incomes
           WHERE substr(date,1,7)=?
           GROUP BY branch, d
           ORDER BY branch, d""",
        (ym,)
    )
    pivot = {}
    for branch, d, amt in cur.fetchall():
        if not branch:
            branch = ""
        pivot.setdefault(branch, {i:0.0 for i in range(1, days+1)})
        pivot[branch][int(d)] = float(amt or 0.0)
    return pivot, days

def export_all(db_path: str, out_path: str, ym: str = "2025-10") -> str:
    if Workbook is None:
        raise RuntimeError("openpyxl yok; lÃ¼tfen kurun.")
    out = pathlib.Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    con = _connect(db_path)
    try:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 1) Ozet_Aylik
        ws = wb.create_sheet("Ozet_Aylik")
        inc, exp, adv, net = _monthly_totals(con, ym)
        ws.append(["Ay", ym])
        ws.append(["Toplam Gelir", inc])
        ws.append(["Toplam Gider", exp])
        ws.append(["Toplam Avans", adv])
        ws.append(["Net", net])

        # 2) Sube_Kirilim
        ws2 = wb.create_sheet("Sube_Kirilim")
        ws2.append(["Åube", "Gelir", "Gider"])
        for branch, inc_b, exp_b in _branch_breakdown(con, ym):
            ws2.append([branch, inc_b, exp_b])

        # 2.5) Sube_Pivot (GÃœNLÃœK GELÄ°R)
        wsp = wb.create_sheet("Sube_Pivot")
        pivot, days = _branch_pivot_incomes(con, ym)
        header = ["Åube"] + [str(d).zfill(2) for d in range(1, days+1)] + ["Toplam"]
        wsp.append(header)
        # rows
        for branch in sorted(pivot.keys(), key=lambda x: (x is None, x.lower() if isinstance(x,str) else "")):
            row = [branch]
            total = 0.0
            for d in range(1, days+1):
                val = round(pivot[branch].get(d, 0.0), 2)
                row.append(val)
                total += val
            row.append(round(total, 2))
            wsp.append(row)

        # 3) Z_Rapor (filled from DB)
        wz = wb.create_sheet("Z_Rapor")
        wz.append(["Tarih", "Kasa Toplam", "POS Toplam", "QR_Dahil"])
        for row in con.execute(
            "SELECT date, cash_total, pos_total, qr_included FROM z_report WHERE substr(date,1,7)=? ORDER BY date", (ym,)
        ):
            wz.append(row)

        # 4) Gelirler
        wg = wb.create_sheet("Gelirler")
        wg.append(["Tarih", "Tutar", "AÃ§Ä±klama", "Åube"])
        for row in con.execute("SELECT date, amount, description, COALESCE(branch,'') FROM incomes ORDER BY date, rowid"):
            wg.append(row)

        # 5) Giderler
        wgd = wb.create_sheet("Giderler")
        wgd.append(["Tarih", "Tutar", "AÃ§Ä±klama", "Åube"])
        for row in con.execute("SELECT date, amount, description, COALESCE(branch,'') FROM expenses ORDER BY date, rowid"):
            wgd.append(row)

        # 6) Avanslar
        wa = wb.create_sheet("Avanslar")
        wa.append(["Personel", "Tarih", "Tutar", "Not"])
        for row in con.execute("SELECT p.name AS person, a.date, a.amount, a.note FROM advances a JOIN personnel p ON p.id=a.personnel_id ORDER BY a.date, a.rowid"):
            wa.append(row)

        wb.save(out)
        return str(out)
    finally:
        con.close()
